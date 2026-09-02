"""Structure-preserving extractors for temporary staged document copies."""

from __future__ import annotations

import html.parser
import math
import re
import shutil
import tempfile
import zipfile
from collections.abc import Iterable
from dataclasses import dataclass, field
from datetime import date, datetime, time
from pathlib import Path
from typing import ClassVar

from defusedxml import ElementTree
from defusedxml.common import DefusedXmlException

from .extraction_errors import ExtractionError

EXTRACTOR_VERSION = "source-units-v4"
EXTRACTOR_VERSION_OVERRIDES = {
    "docx": "source-units-v7",
    "pptx": "source-units-v8",
    "hwpx": "source-units-v10",
    "xlsx": "source-units-v8",
}
MAX_ARCHIVE_MEMBERS = 20_000
MAX_ARCHIVE_EXPANDED_BYTES = 512 * 1024 * 1024
MAX_XML_MEMBER_BYTES = 64 * 1024 * 1024
MAX_UNIT_CHARS = 50_000
MAX_SHEET_ROWS = 100_000
MAX_SHEET_CELLS = 1_000_000
MAX_XLSX_SHEET_STRUCTURE_ITEMS = 100_000
MAX_XLSX_UNITS = 200_000
MAX_XLSX_CONTENT_CHARS = 100_000_000


@dataclass
class UnitDraft:
    unit_type: str
    structure_path: dict
    content: str
    issues: list[dict] = field(default_factory=list)


@dataclass
class ExtractionResult:
    units: list[UnitDraft]
    issues: list[dict] = field(default_factory=list)


def normalize_text(value: str) -> str:
    value = value.replace("\r\n", "\n").replace("\r", "\n")
    value = "\n".join(line.rstrip() for line in value.splitlines())
    return value.strip()


def _bounded_unit(draft: UnitDraft) -> Iterable[UnitDraft]:
    text = normalize_text(draft.content)
    if not text:
        return
    if len(text) <= MAX_UNIT_CHARS:
        yield UnitDraft(draft.unit_type, draft.structure_path, text, draft.issues)
        return
    for chunk_index, start in enumerate(range(0, len(text), MAX_UNIT_CHARS), start=1):
        structure = dict(draft.structure_path)
        structure["chunk"] = chunk_index
        yield UnitDraft(
            draft.unit_type,
            structure,
            text[start : start + MAX_UNIT_CHARS],
            [
                *draft.issues,
                {
                    "code": "unit_split",
                    "message": "Long source unit was split into bounded chunks.",
                },
            ],
        )


def _finish(
    units: Iterable[UnitDraft],
    issues: list[dict] | None = None,
    *,
    preserve_empty: bool = False,
) -> ExtractionResult:
    bounded: list[UnitDraft] = []
    for unit in units:
        if preserve_empty and not normalize_text(unit.content):
            bounded.append(
                UnitDraft(
                    unit.unit_type,
                    unit.structure_path,
                    "",
                    [
                        *unit.issues,
                        {
                            "code": "no_extractable_text",
                            "message": "This structural unit contains no extractable text.",
                        },
                    ],
                )
            )
        else:
            bounded.extend(_bounded_unit(unit))
    result_issues = list(issues or [])
    if not bounded:
        result_issues.append(
            {
                "code": "no_extractable_text",
                "severity": "warning",
                "message": "The extractor found no non-empty text units.",
            }
        )
    return ExtractionResult(units=bounded, issues=result_issues)


def _preflight_zip(path: Path) -> None:
    try:
        with zipfile.ZipFile(path) as archive:
            members = archive.infolist()
            if len(members) > MAX_ARCHIVE_MEMBERS:
                raise ExtractionError(
                    "archive has too many members",
                    details={
                        "member_count": len(members),
                        "limit": MAX_ARCHIVE_MEMBERS,
                    },
                )
            expanded = sum(member.file_size for member in members)
            if expanded > MAX_ARCHIVE_EXPANDED_BYTES:
                raise ExtractionError(
                    "archive expands beyond the extraction limit",
                    details={
                        "expanded_bytes": expanded,
                        "limit": MAX_ARCHIVE_EXPANDED_BYTES,
                    },
                )
    except zipfile.BadZipFile as exc:
        raise ExtractionError(
            "invalid ZIP-based document", details={"error": str(exc)}
        ) from exc


def _safe_archive_xml_root(
    archive: zipfile.ZipFile,
    member_name: str,
):
    member = archive.getinfo(member_name)
    if member.file_size > MAX_XML_MEMBER_BYTES:
        raise ExtractionError(
            "archive XML member exceeds the extraction limit",
            details={
                "reason": "xml_member_too_large",
                "member_bytes": member.file_size,
                "limit": MAX_XML_MEMBER_BYTES,
            },
        )
    with archive.open(member) as stream:
        content = stream.read(MAX_XML_MEMBER_BYTES + 1)
    if len(content) > MAX_XML_MEMBER_BYTES:
        raise ExtractionError(
            "archive XML member exceeds the extraction limit",
            details={
                "reason": "xml_member_too_large",
                "member_bytes": len(content),
                "limit": MAX_XML_MEMBER_BYTES,
            },
        )
    try:
        return ElementTree.fromstring(
            content,
            forbid_dtd=True,
            forbid_entities=True,
            forbid_external=True,
        )
    except DefusedXmlException as exc:
        raise ExtractionError(
            "archive XML contains a forbidden construct",
            details={"reason": "unsafe_xml"},
        ) from exc


def extract_text(path: Path) -> ExtractionResult:
    text = path.read_text(encoding="utf-8", errors="replace")
    paragraphs = re.split(r"\n\s*\n", text)
    units = [
        UnitDraft("paragraph", {"paragraph": index}, paragraph)
        for index, paragraph in enumerate(paragraphs, start=1)
    ]
    return _finish(units)


def extract_markdown(path: Path) -> ExtractionResult:
    text = path.read_text(encoding="utf-8", errors="replace")
    units: list[UnitDraft] = []
    heading_stack: list[str] = []
    paragraph_lines: list[str] = []
    paragraph_start = 1

    def flush_paragraph(end_line: int) -> None:
        nonlocal paragraph_lines
        if paragraph_lines:
            units.append(
                UnitDraft(
                    "paragraph",
                    {
                        "heading_path": list(heading_stack),
                        "line_start": paragraph_start,
                        "line_end": end_line,
                    },
                    "\n".join(paragraph_lines),
                )
            )
            paragraph_lines = []

    for line_number, line in enumerate(text.splitlines(), start=1):
        heading = re.match(r"^(#{1,6})\s+(.+?)\s*#*\s*$", line)
        if heading:
            flush_paragraph(line_number - 1)
            level = len(heading.group(1))
            title = normalize_text(heading.group(2))
            heading_stack[:] = heading_stack[: level - 1]
            heading_stack.append(title)
            units.append(
                UnitDraft(
                    "heading",
                    {
                        "heading_path": list(heading_stack),
                        "level": level,
                        "line_start": line_number,
                        "line_end": line_number,
                    },
                    title,
                )
            )
            paragraph_start = line_number + 1
        elif not line.strip():
            flush_paragraph(line_number - 1)
            paragraph_start = line_number + 1
        else:
            if not paragraph_lines:
                paragraph_start = line_number
            paragraph_lines.append(line)
    flush_paragraph(len(text.splitlines()))
    return _finish(units)


class _HTMLUnitParser(html.parser.HTMLParser):
    BLOCKS: ClassVar[set[str]] = {
        "p",
        "li",
        "blockquote",
        "pre",
        "td",
        "th",
        "tr",
        "figcaption",
        "title",
        "text",
    }
    HEADINGS: ClassVar[set[str]] = {"h1", "h2", "h3", "h4", "h5", "h6"}
    VOID_TAGS: ClassVar[set[str]] = {
        "area",
        "base",
        "br",
        "col",
        "embed",
        "hr",
        "img",
        "input",
        "link",
        "meta",
        "param",
        "source",
        "track",
        "wbr",
    }
    SUPPRESSED_TAGS: ClassVar[set[str]] = {"script", "style", "noscript"}

    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.units: list[UnitDraft] = []
        self._capture_tag: str | None = None
        self._capture_depth = 0
        self._parts: list[str] = []
        self._heading_stack: list[str] = []
        self._counts: dict[str, int] = {}
        self._visible_text: list[str] = []
        self._suppressed_depth = 0

    def handle_starttag(self, tag: str, attrs) -> None:
        tag = tag.lower()
        if tag in self.SUPPRESSED_TAGS:
            self._suppressed_depth += 1
            return
        if self._suppressed_depth:
            return
        if self._capture_tag is not None:
            if tag == "br":
                self._parts.append("\n")
            elif tag not in self.VOID_TAGS:
                self._capture_depth += 1
            return
        if tag in self.BLOCKS | self.HEADINGS:
            self._capture_tag = tag
            self._capture_depth = 1
            self._parts = []

    def handle_data(self, data: str) -> None:
        if self._suppressed_depth:
            return
        if self._capture_tag is not None:
            self._parts.append(data)
        elif data.strip():
            self._visible_text.append(data)

    def handle_endtag(self, tag: str) -> None:
        tag = tag.lower()
        if tag in self.SUPPRESSED_TAGS and self._suppressed_depth:
            self._suppressed_depth -= 1
            return
        if self._suppressed_depth:
            return
        if self._capture_tag is None:
            return
        self._capture_depth -= 1
        if self._capture_depth:
            return
        capture_tag = self._capture_tag
        text = normalize_text(" ".join(self._parts))
        self._capture_tag = None
        self._parts = []
        if not text:
            return
        self._counts[capture_tag] = self._counts.get(capture_tag, 0) + 1
        if capture_tag in self.HEADINGS:
            level = int(capture_tag[1])
            self._heading_stack[:] = self._heading_stack[: level - 1]
            self._heading_stack.append(text)
            unit_type = "heading"
        elif capture_tag in {"td", "th", "tr"}:
            unit_type = "table"
        else:
            unit_type = "paragraph"
        self.units.append(
            UnitDraft(
                unit_type,
                {
                    "tag": capture_tag,
                    "tag_ordinal": self._counts[capture_tag],
                    "heading_path": list(self._heading_stack),
                },
                text,
            )
        )


def extract_html(path: Path) -> ExtractionResult:
    parser = _HTMLUnitParser()
    parser.feed(path.read_text(encoding="utf-8", errors="replace"))
    parser.close()
    if not parser.units:
        fallback = normalize_text(" ".join(parser._visible_text))
        if fallback:
            parser.units.append(
                UnitDraft("document_text", {"scope": "visible_text_fallback"}, fallback)
            )
    return _finish(parser.units)


def _local_name(tag: str) -> str:
    return tag.rsplit("}", 1)[-1]


def extract_hwpx(path: Path) -> ExtractionResult:
    from .hwpx_structure import extract_structured_hwpx

    try:
        return extract_structured_hwpx(path)
    except (zipfile.BadZipFile, ElementTree.ParseError, KeyError) as exc:
        raise ExtractionError("could not parse HWPX structure") from exc


def extract_pdf(
    path: Path, *, page_start: int = 1, max_pages: int | None = None
) -> ExtractionResult:
    try:
        from pypdf import PdfReader
    except ImportError as exc:
        raise ExtractionError("pypdf is not installed") from exc
    try:
        reader = PdfReader(str(path), strict=False)
        if reader.is_encrypted:
            try:
                reader.decrypt("")
            except Exception as exc:
                raise ExtractionError("encrypted PDF cannot be read") from exc
        units = []
        issues = []
        total_pages = len(reader.pages)
        page_end = (
            min(total_pages, page_start - 1 + max_pages)
            if max_pages is not None
            else total_pages
        )
        if not 1 <= page_start <= total_pages + 1:
            raise ExtractionError("PDF page range is invalid")
        for page_number in range(page_start, page_end + 1):
            page = reader.pages[page_number - 1]
            text = page.extract_text() or ""
            if not normalize_text(text):
                issues.append(
                    {
                        "code": "pdf_page_without_text",
                        "severity": "warning",
                        "message": "A PDF page has no extractable text and may require OCR.",
                        "page": page_number,
                    }
                )
            units.append(UnitDraft("page", {"page": page_number}, text))
        if max_pages is not None:
            issues.append(
                {
                    "code": "pdf_page_range_observed",
                    "severity": "info",
                    "message": "The adapter observed this contiguous original page range.",
                    "page_start": page_start,
                    "page_end": page_end,
                    "document_pages": total_pages,
                }
            )
            if page_end < total_pages:
                issues.append(
                    {
                        "code": "pdf_page_range_pending",
                        "severity": "warning",
                        "message": "Further original pages remain for a bounded continuation.",
                        "next_page": page_end + 1,
                        "document_pages": total_pages,
                        "reason": "page_limit",
                    }
                )
    except ExtractionError:
        raise
    except Exception as exc:
        raise ExtractionError(
            "could not extract PDF", details={"error": str(exc)}
        ) from exc
    return _finish(units, issues, preserve_empty=True)


def extract_docx(path: Path) -> ExtractionResult:
    from .office_structure import extract_structured_docx

    return extract_structured_docx(path)


def extract_pptx(path: Path) -> ExtractionResult:
    from .office_structure import extract_structured_pptx

    return extract_structured_pptx(path)


def _cell_value(cell) -> str:
    value = getattr(cell, "value", None)
    if value is None:
        return ""
    if isinstance(value, str):
        return normalize_text(value)
    return str(value)


def _xlsx_scalar(value, data_type: str | None = None) -> dict:
    """Return a JSON-safe scalar without guessing from formatted display text."""

    if data_type == "e":
        return {"kind": "error", "value": str(value)}
    if isinstance(value, bool):
        return {"kind": "boolean", "value": value}
    if isinstance(value, datetime):
        return {"kind": "datetime", "value": value.isoformat()}
    if isinstance(value, date):
        return {"kind": "date", "value": value.isoformat()}
    if isinstance(value, time):
        return {"kind": "time", "value": value.isoformat()}
    if isinstance(value, int):
        return {"kind": "integer", "value": value}
    if isinstance(value, float):
        if not math.isfinite(value):
            raise ExtractionError("XLSX cell contains a non-finite number")
        return {"kind": "number", "value": value}
    if isinstance(value, str):
        return {"kind": "string", "value": value}
    return {
        "kind": "string",
        "value": str(value),
        "coercion": "source_value_string",
    }


def _xlsx_cell_style_metadata(cell) -> tuple[dict, str | None]:
    """Return source style metadata without letting a broken style table drop a cell."""

    metadata: dict = {}
    style_id = getattr(cell, "_style_id", None)
    if isinstance(style_id, int) and not isinstance(style_id, bool) and style_id >= 0:
        metadata["style_id"] = style_id
    try:
        number_format = cell.number_format
    except (AttributeError, IndexError, KeyError, TypeError, ValueError) as exc:
        return metadata, type(exc).__name__
    if isinstance(number_format, str):
        metadata["number_format"] = number_format
        return metadata, None
    return metadata, "InvalidNumberFormat"


def _xlsx_cell_scalar(cell, *, number_format: str | None = None) -> dict:
    value = getattr(cell, "value", None)
    data_type = getattr(cell, "data_type", None)
    if isinstance(value, datetime) and number_format is not None:
        from openpyxl.styles.numbers import is_datetime

        temporal_kind = is_datetime(number_format)
        if temporal_kind == "date":
            return {"kind": "date", "value": value.date().isoformat()}
        if temporal_kind == "time":
            return {"kind": "time", "value": value.time().isoformat()}
    return _xlsx_scalar(value, data_type)


def _xlsx_typed_value(
    cell,
    cached_cell=None,
    *,
    number_format: str | None = None,
) -> dict:
    value = getattr(cell, "value", None)
    data_type = getattr(cell, "data_type", None)
    if data_type != "f":
        return _xlsx_cell_scalar(cell, number_format=number_format)
    cached = getattr(cached_cell, "value", None)
    cached_value = None
    if cached is not None:
        cached_value = _xlsx_cell_scalar(
            cached_cell,
            number_format=number_format,
        )
    return {
        "kind": "formula",
        "formula": str(value),
        "cached_available": cached is not None,
        "cached_value": cached_value,
        "evaluation": "stored_cached_value_only",
    }


_MARKUP_COMPATIBILITY_NAMESPACE = (
    "http://schemas.openxmlformats.org/markup-compatibility/2006"
)
_SPREADSHEETML_NAMESPACE = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_XLSX_STYLES_MEMBER = "xl/styles.xml"


def _resolve_xlsx_style_alternate_content(root) -> int:
    """Select standard fallbacks while preserving the declared style-list order."""

    alternate_tag = f"{{{_MARKUP_COMPATIBILITY_NAMESPACE}}}AlternateContent"
    fallback_tag = f"{{{_MARKUP_COMPATIBILITY_NAMESPACE}}}Fallback"
    resolved = 0
    while True:
        pass_resolved = 0
        for parent in list(root.iter()):
            for index, child in enumerate(list(parent)):
                if child.tag != alternate_tag:
                    continue
                fallback = next(
                    (branch for branch in child if branch.tag == fallback_tag),
                    None,
                )
                if fallback is None:
                    continue
                replacements = list(fallback)
                parent.remove(child)
                for offset, replacement in enumerate(replacements):
                    parent.insert(index + offset, replacement)
                pass_resolved += 1
        resolved += pass_resolved
        if not pass_resolved:
            return resolved


def _xlsx_compatible_styles(styles: bytes) -> tuple[bytes, dict[str, int]]:
    """Make source-declared compatibility styles readable in a temporary copy."""

    try:
        root = ElementTree.fromstring(styles)
    except (DefusedXmlException, ElementTree.ParseError) as exc:
        raise ExtractionError("could not parse XLSX styles XML") from exc
    resolved = _resolve_xlsx_style_alternate_content(root)
    fonts = root.find(f"{{{_SPREADSHEETML_NAMESPACE}}}fonts")
    family_tag = f"{{{_SPREADSHEETML_NAMESPACE}}}family"
    removed = 0
    if fonts is not None:
        for font in list(fonts):
            for child in list(font):
                if child.tag != family_tag:
                    continue
                raw_value = child.attrib.get("val")
                try:
                    value = int(raw_value) if raw_value is not None else None
                except ValueError:
                    continue
                if value is not None and not 0 <= value <= 14:
                    font.remove(child)
                    removed += 1
    changes = {
        "alternate_content_fallbacks": resolved,
        "invalid_font_families": removed,
    }
    if not resolved and not removed:
        return styles, changes
    return ElementTree.tostring(
        root,
        encoding="utf-8",
        xml_declaration=True,
    ), changes


def _read_xlsx_compatible_styles(path: Path) -> tuple[bytes | None, dict[str, int]]:
    with zipfile.ZipFile(path) as package:
        try:
            member = package.getinfo(_XLSX_STYLES_MEMBER)
        except KeyError:
            return None, {
                "alternate_content_fallbacks": 0,
                "invalid_font_families": 0,
            }
        if member.file_size > MAX_XML_MEMBER_BYTES:
            raise ExtractionError(
                "XLSX styles XML exceeds the extraction limit",
                details={
                    "member_bytes": member.file_size,
                    "limit": MAX_XML_MEMBER_BYTES,
                },
            )
        styles, changes = _xlsx_compatible_styles(package.read(member))
    if any(changes.values()):
        return styles, changes
    return None, changes


def _write_xlsx_with_compatible_styles(
    source: Path,
    destination: Path,
    styles: bytes,
) -> None:
    """Write a temporary package containing already-normalized style XML."""

    with (
        zipfile.ZipFile(source) as package,
        zipfile.ZipFile(destination, "w", allowZip64=True) as normalized,
    ):
        normalized.comment = package.comment
        for member in package.infolist():
            if member.filename == _XLSX_STYLES_MEMBER:
                normalized.writestr(member, styles)
                continue
            with (
                package.open(member) as source_member,
                normalized.open(member, "w", force_zip64=True) as destination_member,
            ):
                shutil.copyfileobj(
                    source_member, destination_member, length=1024 * 1024
                )


def _open_xlsx_workbook(load_workbook, path: Path, *, data_only: bool):
    package = path.open("rb")
    try:
        workbook = load_workbook(
            filename=package,
            read_only=True,
            data_only=data_only,
            keep_links=False,
        )
    except Exception:
        package.close()
        raise
    return package, workbook


def _xlsx_sheet_structure(
    archive: zipfile.ZipFile,
    worksheet_path: str,
) -> tuple[dict, list[dict]]:
    """Read bounded, source-declared sheet structure omitted by read-only OpenPyXL."""

    from openpyxl.utils.cell import range_boundaries

    structure = {
        "merged_ranges": [],
        "hidden_rows": [],
        "hidden_column_ranges": [],
    }
    issues: list[dict] = []
    try:
        member = archive.getinfo(worksheet_path)
    except KeyError:
        return structure, [
            {
                "code": "xlsx_sheet_structure_partial",
                "severity": "warning",
                "message": "The worksheet XML part could not be located.",
                "details": {"worksheet_part": worksheet_path},
            }
        ]
    if member.file_size > MAX_XML_MEMBER_BYTES:
        return structure, [
            {
                "code": "xlsx_sheet_structure_partial",
                "severity": "warning",
                "message": (
                    "Merged-cell and hidden-range metadata exceeded the bounded XML "
                    "inspection limit."
                ),
                "details": {
                    "worksheet_part": worksheet_path,
                    "member_bytes": member.file_size,
                    "limit": MAX_XML_MEMBER_BYTES,
                },
            }
        ]
    root = _safe_archive_xml_root(archive, worksheet_path)

    item_count = 0
    invalid_ranges = 0
    for node in root.iter():
        name = _local_name(node.tag)
        if name == "dimension" and node.get("ref"):
            structure["declared_dimension"] = node.get("ref")
        elif name == "mergeCell" and node.get("ref"):
            if item_count >= MAX_XLSX_SHEET_STRUCTURE_ITEMS:
                break
            reference = node.get("ref")
            try:
                min_col, min_row, max_col, max_row = range_boundaries(reference)
                if min_col < 1 or min_row < 1 or max_col < min_col or max_row < min_row:
                    raise ValueError("invalid merged range")
            except (TypeError, ValueError):
                invalid_ranges += 1
                continue
            structure["merged_ranges"].append(
                {
                    "range": reference,
                    "origin": {"row": min_row, "col": min_col},
                    "row_span": max_row - min_row + 1,
                    "col_span": max_col - min_col + 1,
                }
            )
            item_count += 1
        elif name == "row" and node.get("hidden") in {"1", "true"}:
            if item_count >= MAX_XLSX_SHEET_STRUCTURE_ITEMS:
                break
            try:
                row = int(node.get("r", ""))
            except ValueError:
                invalid_ranges += 1
                continue
            if row < 1:
                invalid_ranges += 1
                continue
            structure["hidden_rows"].append(row)
            item_count += 1
        elif name == "col" and node.get("hidden") in {"1", "true"}:
            if item_count >= MAX_XLSX_SHEET_STRUCTURE_ITEMS:
                break
            try:
                minimum = int(node.get("min", ""))
                maximum = int(node.get("max", ""))
            except ValueError:
                invalid_ranges += 1
                continue
            if minimum < 1 or maximum < minimum:
                invalid_ranges += 1
                continue
            structure["hidden_column_ranges"].append({"min_col": minimum, "max_col": maximum})
            item_count += 1

    if item_count >= MAX_XLSX_SHEET_STRUCTURE_ITEMS:
        issues.append(
            {
                "code": "xlsx_sheet_structure_partial",
                "severity": "warning",
                "message": "Worksheet structure metadata reached its configured item limit.",
                "details": {
                    "worksheet_part": worksheet_path,
                    "limit": MAX_XLSX_SHEET_STRUCTURE_ITEMS,
                },
            }
        )
    if invalid_ranges:
        issues.append(
            {
                "code": "xlsx_sheet_structure_partial",
                "severity": "warning",
                "message": "Some worksheet range metadata was invalid and was not projected.",
                "details": {
                    "worksheet_part": worksheet_path,
                    "occurrences": invalid_ranges,
                },
            }
        )
    return structure, issues


def extract_xlsx(path: Path) -> ExtractionResult:
    _preflight_zip(path)
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ExtractionError("openpyxl is not installed") from exc
    units: list[UnitDraft] = []
    issues: list[dict] = []
    cells_seen = 0
    content_chars = 0
    package = cached_package = None
    workbook = cached_workbook = None
    temporary = None
    workbook_path = path
    try:
        compatible_styles, style_changes = _read_xlsx_compatible_styles(path)
        if compatible_styles is not None:
            temporary = tempfile.TemporaryDirectory(prefix="document-files-xlsx-")
            normalized_path = Path(temporary.name) / "normalized.xlsx"
            _write_xlsx_with_compatible_styles(
                path,
                normalized_path,
                compatible_styles,
            )
            workbook_path = normalized_path
            if style_changes["alternate_content_fallbacks"]:
                issues.append(
                    {
                        "code": "xlsx_alternate_content_resolved",
                        "severity": "info",
                        "message": (
                            "Standard SpreadsheetML style fallbacks were selected in "
                            "a temporary read-only copy."
                        ),
                        "details": {
                            "occurrences": style_changes[
                                "alternate_content_fallbacks"
                            ]
                        },
                    }
                )
            if style_changes["invalid_font_families"]:
                issues.append(
                    {
                        "code": "xlsx_invalid_font_family_ignored",
                        "severity": "info",
                        "message": (
                            "Invalid XLSX font-family metadata was ignored in a "
                            "temporary read-only copy."
                        ),
                        "details": {
                            "removed_elements": style_changes[
                                "invalid_font_families"
                            ]
                        },
                    }
                )
        package, workbook = _open_xlsx_workbook(
            load_workbook,
            workbook_path,
            data_only=False,
        )
        cached_package, cached_workbook = _open_xlsx_workbook(
            load_workbook,
            workbook_path,
            data_only=True,
        )
        try:
            if len(workbook.worksheets) != len(cached_workbook.worksheets):
                raise ExtractionError("XLSX formula and cached-value views disagree")
            missing_formula_cache = 0
            unresolved_cell_styles = 0
            unresolved_cell_style_samples: list[dict] = []
            unresolved_cell_style_errors: dict[str, int] = {}
            workbook_limit_kind = None
            with zipfile.ZipFile(workbook_path) as structure_archive:
                for sheet_index, (sheet, cached_sheet) in enumerate(
                    zip(workbook.worksheets, cached_workbook.worksheets, strict=True),
                    start=1,
                ):
                    if len(units) >= MAX_XLSX_UNITS:
                        workbook_limit_kind = "units"
                        issues.append(
                            {
                                "code": "sheet_limit_reached",
                                "severity": "warning",
                                "message": (
                                    "Workbook extraction stopped at a configured output limit."
                                ),
                                "sheet": sheet.title,
                                "details": {
                                    "limit": MAX_XLSX_UNITS,
                                    "kind": workbook_limit_kind,
                                },
                            }
                        )
                        break
                    if sheet.title != cached_sheet.title:
                        raise ExtractionError("XLSX worksheet order is inconsistent")
                    worksheet_path = getattr(sheet, "_worksheet_path", "")
                    sheet_structure, sheet_issues = _xlsx_sheet_structure(
                        structure_archive,
                        worksheet_path,
                    )
                    issues.extend(sheet_issues)
                    units.append(
                        UnitDraft(
                            "sheet",
                            {
                                "sheet": sheet.title,
                                "sheet_index": sheet_index,
                                "sheet_state": sheet.sheet_state,
                                "max_row": sheet.max_row,
                                "max_col": sheet.max_column,
                                **sheet_structure,
                            },
                            "",
                        )
                    )
                    merge_origins = {
                        (item["origin"]["row"], item["origin"]["col"]): item
                        for item in sheet_structure["merged_ranges"]
                    }
                    cached_rows = iter(cached_sheet.iter_rows())
                    for row_index, row in enumerate(sheet.iter_rows(), start=1):
                        if row_index > MAX_SHEET_ROWS:
                            issues.append(
                                {
                                    "code": "sheet_limit_reached",
                                    "severity": "warning",
                                    "message": (
                                        "Worksheet extraction stopped at the configured "
                                        "row limit."
                                    ),
                                    "sheet": sheet.title,
                                    "details": {
                                        "limit": MAX_SHEET_ROWS,
                                        "kind": "rows",
                                    },
                                }
                            )
                            break
                        cached_row = next(cached_rows, ())
                        for col_index, cell in enumerate(row, start=1):
                            if cells_seen >= MAX_SHEET_CELLS:
                                workbook_limit_kind = "cells"
                                break
                            cells_seen += 1
                            value = getattr(cell, "value", None)
                            if value is None:
                                continue
                            cached_cell = (
                                cached_row[col_index - 1]
                                if col_index <= len(cached_row)
                                else None
                            )
                            style_metadata, style_error = _xlsx_cell_style_metadata(cell)
                            if style_error is not None:
                                unresolved_cell_styles += 1
                                unresolved_cell_style_errors[style_error] = (
                                    unresolved_cell_style_errors.get(style_error, 0) + 1
                                )
                                if len(unresolved_cell_style_samples) < 20:
                                    unresolved_cell_style_samples.append(
                                        {
                                            "sheet": sheet.title,
                                            "coordinate": getattr(
                                                cell,
                                                "coordinate",
                                                None,
                                            ),
                                            "style_id": style_metadata.get("style_id"),
                                        }
                                    )
                            typed_value = _xlsx_typed_value(
                                cell,
                                cached_cell,
                                number_format=style_metadata.get("number_format"),
                            )
                            if (
                                typed_value["kind"] == "formula"
                                and not typed_value["cached_available"]
                            ):
                                missing_formula_cache += 1
                            coordinate = getattr(cell, "coordinate", None)
                            if not coordinate:
                                raise ExtractionError(
                                    "XLSX non-empty cell is missing its coordinate"
                                )
                            cell_content = f"{coordinate}={_cell_value(cell)}"
                            if len(units) >= MAX_XLSX_UNITS:
                                workbook_limit_kind = "units"
                                break
                            if content_chars + len(cell_content) > MAX_XLSX_CONTENT_CHARS:
                                workbook_limit_kind = "content_chars"
                                break
                            location = {
                                "sheet": sheet.title,
                                "sheet_index": sheet_index,
                                "sheet_state": sheet.sheet_state,
                                "row": row_index,
                                "col": col_index,
                                "coordinate": coordinate,
                                "value": typed_value,
                                **style_metadata,
                            }
                            if merge := merge_origins.get((row_index, col_index)):
                                location.update(
                                    {
                                        "merged_range": merge["range"],
                                        "row_span": merge["row_span"],
                                        "col_span": merge["col_span"],
                                    }
                                )
                            units.append(
                                UnitDraft(
                                    "sheet_cell",
                                    location,
                                    cell_content,
                                )
                            )
                            content_chars += len(cell_content)
                        if workbook_limit_kind:
                            break
                    if workbook_limit_kind:
                        limits = {
                            "cells": MAX_SHEET_CELLS,
                            "units": MAX_XLSX_UNITS,
                            "content_chars": MAX_XLSX_CONTENT_CHARS,
                        }
                        issues.append(
                            {
                                "code": "sheet_limit_reached",
                                "severity": "warning",
                                "message": (
                                    "Workbook extraction stopped at a configured output limit."
                                ),
                                "sheet": sheet.title,
                                "details": {
                                    "limit": limits[workbook_limit_kind],
                                    "kind": workbook_limit_kind,
                                },
                            }
                        )
                        break
            if missing_formula_cache:
                issues.append(
                    {
                        "code": "xlsx_formula_without_cached_value",
                        "severity": "info",
                        "impact": "observation",
                        "coverage_dimensions": [],
                        "message": (
                            "Some formulas have no stored cached result; their expressions "
                            "were retained without evaluation."
                        ),
                        "details": {"occurrences": missing_formula_cache},
                    }
                )
            if unresolved_cell_styles:
                issues.append(
                    {
                        "code": "xlsx_cell_style_partial",
                        "severity": "warning",
                        "impact": "structure_gap",
                        "coverage_dimensions": ["structure"],
                        "message": (
                            "Some cell style metadata could not be resolved; source "
                            "values and raw style references were retained."
                        ),
                        "details": {
                            "occurrences": unresolved_cell_styles,
                            "error_types": unresolved_cell_style_errors,
                            "samples": unresolved_cell_style_samples,
                        },
                    }
                )
        finally:
            workbook.close()
            package.close()
            cached_workbook.close()
            cached_package.close()
            workbook = None
            package = None
            cached_workbook = None
            cached_package = None
    except Exception as exc:
        raise ExtractionError("could not extract XLSX", details={"error": str(exc)}) from exc
    finally:
        if workbook is not None:
            workbook.close()
        if package is not None:
            package.close()
        if cached_workbook is not None:
            cached_workbook.close()
        if cached_package is not None:
            cached_package.close()
        if temporary is not None:
            temporary.cleanup()
    if not any(unit.content for unit in units):
        issues.append(
            {
                "code": "no_extractable_text",
                "severity": "warning",
                "message": "The workbook contains structure but no non-empty cell values.",
            }
        )
    return ExtractionResult(units=units, issues=issues)


EXTRACTORS = {
    "text": extract_text,
    "markdown": extract_markdown,
    "html": extract_html,
    "hwpx": extract_hwpx,
    "pdf": extract_pdf,
    "docx": extract_docx,
    "pptx": extract_pptx,
    "xlsx": extract_xlsx,
}


def extract(path: Path, adapter: str) -> ExtractionResult:
    extractor = EXTRACTORS.get(adapter)
    if extractor is None:
        raise ExtractionError(
            "no extractor is registered", details={"adapter": adapter}
        )
    return extractor(path)
