from __future__ import annotations

import hashlib
import struct
import zipfile
from datetime import date
from pathlib import Path

from document_files.engine import extract_structure
from document_files.extraction_protocol import run_builtin_extraction
from document_files.hwp_structure import SectionStructure, doc_info_properties
from document_files.structured_extraction import project_structured_extraction
from openpyxl import Workbook


def _digest(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def test_xlsx_structure_and_typed_values_are_source_addressed(tmp_path: Path) -> None:
    path = tmp_path / "values.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "데이터"
    sheet.merge_cells("A1:B1")
    sheet["A1"] = "항목"
    sheet["A2"] = "수량"
    sheet["B2"] = 3
    sheet["C2"] = True
    sheet["D2"] = date(2026, 9, 2)
    sheet["E2"] = "=B2*2"
    sheet.row_dimensions[3].hidden = True
    sheet.column_dimensions["F"].hidden = True
    hidden = workbook.create_sheet("비공개")
    hidden.sheet_state = "hidden"
    hidden["A1"] = "숨김값"
    workbook.save(path)
    before = _digest(path)

    result = extract_structure(path, max_units=50)

    assert _digest(path) == before
    assert result["schemaVersion"] == "document-files.structured-extraction.v1"
    assert result["semanticPolicy"] == {
        "sourceDeclaredOnly": True,
        "adjacentCellInference": False,
        "formulaEvaluation": False,
        "indexIdentityAssigned": False,
    }
    assert result["summary"]["unitTypes"] == {"sheet": 2, "sheet_cell": 7}
    units = result["units"]
    data_sheet = next(
        unit
        for unit in units
        if unit["sourceUnitType"] == "sheet" and unit["semantic"]["sheet"]["name"] == "데이터"
    )
    assert data_sheet["semantic"]["sheet"]["mergedRanges"] == [
        {
            "range": "A1:B1",
            "origin": {"row": 1, "column": 1},
            "rowSpan": 1,
            "columnSpan": 2,
        }
    ]
    assert data_sheet["semantic"]["sheet"]["hiddenRows"] == [3]
    assert data_sheet["semantic"]["sheet"]["hiddenColumnRanges"] == [
        {"minColumn": 6, "maxColumn": 6}
    ]

    cells = {
        unit["semantic"]["cell"]["coordinate"]: unit["semantic"]
        for unit in units
        if unit["sourceUnitType"] == "sheet_cell" and unit["semantic"]["sheet"]["name"] == "데이터"
    }
    assert cells["B2"]["value"] == {"kind": "integer", "value": 3}
    assert cells["C2"]["value"] == {"kind": "boolean", "value": True}
    assert cells["D2"]["value"] == {"kind": "date", "value": "2026-09-02"}
    assert cells["D2"]["cell"]["numberFormat"] == "yyyy-mm-dd"
    assert cells["D2"]["cell"]["styleId"] > 0
    assert cells["E2"]["value"] == {
        "kind": "formula",
        "formula": "=B2*2",
        "evaluation": "stored_cached_value_only",
        "cachedAvailable": False,
        "cachedValue": None,
    }
    assert result["summary"]["issueCounts"] == {"xlsx_formula_without_cached_value": 1}


def test_structured_extraction_is_paged_and_can_omit_text(tmp_path: Path) -> None:
    path = tmp_path / "source.md"
    path.write_text("# 제목\n\n첫 문단\n\n둘째 문단", encoding="utf-8")

    first = extract_structure(path, max_units=2, include_text=False)
    second = extract_structure(
        path,
        unit_offset=first["unitPage"]["nextOffset"],
        max_units=2,
    )

    assert first["unitPage"] == {
        "offset": 0,
        "limit": 2,
        "returned": 2,
        "total": 3,
        "hasMore": True,
        "nextOffset": 2,
        "textIncluded": False,
    }
    assert all("text" not in unit for unit in first["units"])
    assert second["unitPage"]["hasMore"] is False
    assert [unit["text"] for unit in second["units"]] == ["둘째 문단"]
    assert first["manifestHash"] == second["manifestHash"]


def test_hwp_image_bullet_keeps_its_embedded_source_reference() -> None:
    bullet = bytearray(25)
    bullet[12:14] = "◆".encode("utf-16-le")
    struct.pack_into("<I", bullet, 14, 1)
    bullet[18:22] = bytes((0, 0, 0, 1))
    shape = bytearray(32)
    struct.pack_into("<I", shape, 0, 3 << 23)
    struct.pack_into("<H", shape, 30, 1)
    shapes, styles = doc_info_properties(
        [(1, 0x18, 0, bytes(bullet)), (2, 0x19, 0, bytes(shape))]
    )
    reader = SectionStructure(
        1,
        "Section0",
        shapes,
        styles,
        images=[{"bindata_record": 3, "image_parts": ["BinData/BIN0001.png"]}],
    )
    paragraph = bytearray(24)
    struct.pack_into("<H", paragraph, 8, 0)
    reader.observe(3, 0x42, 0, bytes(paragraph))

    reader.text(4, 1, 1, "항목")
    units, issues = reader.finish()

    marker = units[0]["structure_path"]["marker_image"]
    assert marker["binary_item_ref"] == 1
    assert marker["image_parts"] == ["BinData/BIN0001.png"]
    assert marker["fallback_text"] == "◆"
    assert "hwp_list_marker_partial" not in {issue["code"] for issue in issues}


def test_hwpx_image_bullet_projects_a_format_neutral_semantic_marker(
    tmp_path: Path,
) -> None:
    path = tmp_path / "image-bullet.hwpx"
    with zipfile.ZipFile(path, "w") as archive:
        archive.writestr(
            "Contents/header.xml",
            """<header><paraPr id="1"><heading type="BULLET" level="0" idRef="7"/>
            </paraPr><bullet id="7" char="◆" useImage="1"><img
            binaryItemIDRef="image7" bright="0" contrast="0" effect="REAL_PIC"
            alpha="0"/></bullet></header>""",
        )
        archive.writestr(
            "Contents/content.hpf",
            """<package><manifest><item id="body" href="section0.xml"/>
            <item id="image7" href="BinData/image7.png" isEmbeded="1"/>
            </manifest><spine><itemref idref="body"/></spine></package>""",
        )
        archive.writestr(
            "Contents/section0.xml",
            '<sec><p paraPrIDRef="1"><run><t>항목</t></run></p></sec>',
        )
        archive.writestr("BinData/image7.png", b"source image bytes")

    envelope = run_builtin_extraction(path, "hwpx")
    projected = project_structured_extraction(
        envelope,
        source_format="hwpx",
        unit_offset=0,
        max_units=10,
        include_text=True,
    )

    assert "hwpx_list_marker_partial" not in projected["summary"]["issueCounts"]
    assert projected["units"][0]["semantic"]["list"]["marker"] == {
        "kind": "image",
        "basis": "source_image_reference",
        "sourceRef": "image7",
        "sourceParts": ["BinData/image7.png"],
        "fallbackText": "◆",
        "rendering": {
            "brightness": "0",
            "contrast": "0",
            "effect": "REAL_PIC",
            "alpha": "0",
        },
    }
