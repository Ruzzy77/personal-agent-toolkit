"""Structure-preserving extractors for temporary staged document copies."""

from __future__ import annotations

import html.parser
import re
import shutil
import tempfile
import zipfile
from collections.abc import Iterable
from dataclasses import dataclass, field
from pathlib import Path
from typing import ClassVar

from defusedxml import ElementTree
from defusedxml.common import DefusedXmlException

from .errors import ExtractionError

EXTRACTOR_VERSION = "source-units-v4"
EXTRACTOR_VERSION_OVERRIDES = {
    "docx": "source-units-v7",
    "pptx": "source-units-v8",
    "hwpx": "source-units-v9",
    "xlsx": "source-units-v5",
}
MAX_ARCHIVE_MEMBERS = 20_000
MAX_ARCHIVE_EXPANDED_BYTES = 512 * 1024 * 1024
MAX_XML_MEMBER_BYTES = 64 * 1024 * 1024
MAX_UNIT_CHARS = 50_000
MAX_SHEET_ROWS = 100_000
MAX_SHEET_CELLS = 1_000_000


@dataclass
class UnitDraft:
    unit_type: str
    structure_path: dict
    content: str
    issues: list[dict] = field(default_factory=list)


@dataclass
class ExtractionResult:
    units: list[UnitDraft]
    issues: list[dict] = field(default_factory=list)


def normalize_text(value: str) -> str:
    value = value.replace("\r\n", "\n").replace("\r", "\n")
    value = "\n".join(line.rstrip() for line in value.splitlines())
    return value.strip()


def _bounded_unit(draft: UnitDraft) -> Iterable[UnitDraft]:
    text = normalize_text(draft.content)
    if not text:
        return
    if len(text) <= MAX_UNIT_CHARS:
        yield UnitDraft(draft.unit_type, draft.structure_path, text, draft.issues)
        return
    for chunk_index, start in enumerate(range(0, len(text), MAX_UNIT_CHARS), start=1):
        structure = dict(draft.structure_path)
        structure["chunk"] = chunk_index
        yield UnitDraft(
            draft.unit_type,
            structure,
            text[start : start + MAX_UNIT_CHARS],
            [
                *draft.issues,
                {
                    "code": "unit_split",
                    "message": "Long source unit was split into bounded chunks.",
                },
            ],
        )


def _finish(
    units: Iterable[UnitDraft],
    issues: list[dict] | None = None,
    *,
    preserve_empty: bool = False,
) -> ExtractionResult:
    bounded: list[UnitDraft] = []
    for unit in units:
        if preserve_empty and not normalize_text(unit.content):
            bounded.append(
                UnitDraft(
                    unit.unit_type,
                    unit.structure_path,
                    "",
                    [
                        *unit.issues,
                        {
                            "code": "no_extractable_text",
                            "message": "This structural unit contains no extractable text.",
                        },
                    ],
                )
            )
        else:
            bounded.extend(_bounded_unit(unit))
    result_issues = list(issues or [])
    if not bounded:
        result_issues.append(
            {
                "code": "no_extractable_text",
                "severity": "warning",
                "message": "The extractor found no non-empty text units.",
            }
        )
    return ExtractionResult(units=bounded, issues=result_issues)


def _preflight_zip(path: Path) -> None:
    try:
        with zipfile.ZipFile(path) as archive:
            members = archive.infolist()
            if len(members) > MAX_ARCHIVE_MEMBERS:
                raise ExtractionError(
                    "archive has too many members",
                    details={
                        "member_count": len(members),
                        "limit": MAX_ARCHIVE_MEMBERS,
                    },
                )
            expanded = sum(member.file_size for member in members)
            if expanded > MAX_ARCHIVE_EXPANDED_BYTES:
                raise ExtractionError(
                    "archive expands beyond the extraction limit",
                    details={
                        "expanded_bytes": expanded,
                        "limit": MAX_ARCHIVE_EXPANDED_BYTES,
                    },
                )
    except zipfile.BadZipFile as exc:
        raise ExtractionError(
            "invalid ZIP-based document", details={"error": str(exc)}
        ) from exc


def _safe_archive_xml_root(
    archive: zipfile.ZipFile,
    member_name: str,
):
    member = archive.getinfo(member_name)
    if member.file_size > MAX_XML_MEMBER_BYTES:
        raise ExtractionError(
            "archive XML member exceeds the extraction limit",
            details={
                "reason": "xml_member_too_large",
                "member_bytes": member.file_size,
                "limit": MAX_XML_MEMBER_BYTES,
            },
        )
    with archive.open(member) as stream:
        content = stream.read(MAX_XML_MEMBER_BYTES + 1)
    if len(content) > MAX_XML_MEMBER_BYTES:
        raise ExtractionError(
            "archive XML member exceeds the extraction limit",
            details={
                "reason": "xml_member_too_large",
                "member_bytes": len(content),
                "limit": MAX_XML_MEMBER_BYTES,
            },
        )
    try:
        return ElementTree.fromstring(
            content,
            forbid_dtd=True,
            forbid_entities=True,
            forbid_external=True,
        )
    except DefusedXmlException as exc:
        raise ExtractionError(
            "archive XML contains a forbidden construct",
            details={"reason": "unsafe_xml"},
        ) from exc


def extract_text(path: Path) -> ExtractionResult:
    text = path.read_text(encoding="utf-8", errors="replace")
    paragraphs = re.split(r"\n\s*\n", text)
    units = [
        UnitDraft("paragraph", {"paragraph": index}, paragraph)
        for index, paragraph in enumerate(paragraphs, start=1)
    ]
    return _finish(units)


def extract_markdown(path: Path) -> ExtractionResult:
    text = path.read_text(encoding="utf-8", errors="replace")
    units: list[UnitDraft] = []
    heading_stack: list[str] = []
    paragraph_lines: list[str] = []
    paragraph_start = 1

    def flush_paragraph(end_line: int) -> None:
        nonlocal paragraph_lines
        if paragraph_lines:
            units.append(
                UnitDraft(
                    "paragraph",
                    {
                        "heading_path": list(heading_stack),
                        "line_start": paragraph_start,
                        "line_end": end_line,
                    },
                    "\n".join(paragraph_lines),
                )
            )
            paragraph_lines = []

    for line_number, line in enumerate(text.splitlines(), start=1):
        heading = re.match(r"^(#{1,6})\s+(.+?)\s*#*\s*$", line)
        if heading:
            flush_paragraph(line_number - 1)
            level = len(heading.group(1))
            title = normalize_text(heading.group(2))
            heading_stack[:] = heading_stack[: level - 1]
            heading_stack.append(title)
            units.append(
                UnitDraft(
                    "heading",
                    {
                        "heading_path": list(heading_stack),
                        "level": level,
                        "line_start": line_number,
                        "line_end": line_number,
                    },
                    title,
                )
            )
            paragraph_start = line_number + 1
        elif not line.strip():
            flush_paragraph(line_number - 1)
            paragraph_start = line_number + 1
        else:
            if not paragraph_lines:
                paragraph_start = line_number
            paragraph_lines.append(line)
    flush_paragraph(len(text.splitlines()))
    return _finish(units)


class _HTMLUnitParser(html.parser.HTMLParser):
    BLOCKS: ClassVar[set[str]] = {
        "p",
        "li",
        "blockquote",
        "pre",
        "td",
        "th",
        "tr",
        "figcaption",
        "title",
        "text",
    }
    HEADINGS: ClassVar[set[str]] = {"h1", "h2", "h3", "h4", "h5", "h6"}
    VOID_TAGS: ClassVar[set[str]] = {
        "area",
        "base",
        "br",
        "col",
        "embed",
        "hr",
        "img",
        "input",
        "link",
        "meta",
        "param",
        "source",
        "track",
        "wbr",
    }
    SUPPRESSED_TAGS: ClassVar[set[str]] = {"script", "style", "noscript"}

    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.units: list[UnitDraft] = []
        self._capture_tag: str | None = None
        self._capture_depth = 0
        self._parts: list[str] = []
        self._heading_stack: list[str] = []
        self._counts: dict[str, int] = {}
        self._visible_text: list[str] = []
        self._suppressed_depth = 0

    def handle_starttag(self, tag: str, attrs) -> None:
        tag = tag.lower()
        if tag in self.SUPPRESSED_TAGS:
            self._suppressed_depth += 1
            return
        if self._suppressed_depth:
            return
        if self._capture_tag is not None:
            if tag == "br":
                self._parts.append("\n")
            elif tag not in self.VOID_TAGS:
                self._capture_depth += 1
            return
        if tag in self.BLOCKS | self.HEADINGS:
            self._capture_tag = tag
            self._capture_depth = 1
            self._parts = []

    def handle_data(self, data: str) -> None:
        if self._suppressed_depth:
            return
        if self._capture_tag is not None:
            self._parts.append(data)
        elif data.strip():
            self._visible_text.append(data)

    def handle_endtag(self, tag: str) -> None:
        tag = tag.lower()
        if tag in self.SUPPRESSED_TAGS and self._suppressed_depth:
            self._suppressed_depth -= 1
            return
        if self._suppressed_depth:
            return
        if self._capture_tag is None:
            return
        self._capture_depth -= 1
        if self._capture_depth:
            return
        capture_tag = self._capture_tag
        text = normalize_text(" ".join(self._parts))
        self._capture_tag = None
        self._parts = []
        if not text:
            return
        self._counts[capture_tag] = self._counts.get(capture_tag, 0) + 1
        if capture_tag in self.HEADINGS:
            level = int(capture_tag[1])
            self._heading_stack[:] = self._heading_stack[: level - 1]
            self._heading_stack.append(text)
            unit_type = "heading"
        elif capture_tag in {"td", "th", "tr"}:
            unit_type = "table"
        else:
            unit_type = "paragraph"
        self.units.append(
            UnitDraft(
                unit_type,
                {
                    "tag": capture_tag,
                    "tag_ordinal": self._counts[capture_tag],
                    "heading_path": list(self._heading_stack),
                },
                text,
            )
        )


def extract_html(path: Path) -> ExtractionResult:
    parser = _HTMLUnitParser()
    parser.feed(path.read_text(encoding="utf-8", errors="replace"))
    parser.close()
    if not parser.units:
        fallback = normalize_text(" ".join(parser._visible_text))
        if fallback:
            parser.units.append(
                UnitDraft("document_text", {"scope": "visible_text_fallback"}, fallback)
            )
    return _finish(parser.units)


def _local_name(tag: str) -> str:
    return tag.rsplit("}", 1)[-1]


def extract_hwpx(path: Path) -> ExtractionResult:
    from .hwpx_structure import extract_structured_hwpx

    try:
        return extract_structured_hwpx(path)
    except (zipfile.BadZipFile, ElementTree.ParseError, KeyError) as exc:
        raise ExtractionError("could not parse HWPX structure") from exc


def extract_pdf(
    path: Path, *, page_start: int = 1, max_pages: int | None = None
) -> ExtractionResult:
    try:
        from pypdf import PdfReader
    except ImportError as exc:
        raise ExtractionError("pypdf is not installed") from exc
    try:
        reader = PdfReader(str(path), strict=False)
        if reader.is_encrypted:
            try:
                reader.decrypt("")
            except Exception as exc:
                raise ExtractionError("encrypted PDF cannot be read") from exc
        units = []
        issues = []
        total_pages = len(reader.pages)
        page_end = (
            min(total_pages, page_start - 1 + max_pages)
            if max_pages is not None
            else total_pages
        )
        if not 1 <= page_start <= total_pages + 1:
            raise ExtractionError("PDF page range is invalid")
        for page_number in range(page_start, page_end + 1):
            page = reader.pages[page_number - 1]
            text = page.extract_text() or ""
            if not normalize_text(text):
                issues.append(
                    {
                        "code": "pdf_page_without_text",
                        "severity": "warning",
                        "message": "A PDF page has no extractable text and may require OCR.",
                        "page": page_number,
                    }
                )
            units.append(UnitDraft("page", {"page": page_number}, text))
        if max_pages is not None:
            issues.append(
                {
                    "code": "pdf_page_range_observed",
                    "severity": "info",
                    "message": "The adapter observed this contiguous original page range.",
                    "page_start": page_start,
                    "page_end": page_end,
                    "document_pages": total_pages,
                }
            )
            if page_end < total_pages:
                issues.append(
                    {
                        "code": "pdf_page_range_pending",
                        "severity": "warning",
                        "message": "Further original pages remain for a bounded continuation.",
                        "next_page": page_end + 1,
                        "document_pages": total_pages,
                        "reason": "page_limit",
                    }
                )
    except ExtractionError:
        raise
    except Exception as exc:
        raise ExtractionError(
            "could not extract PDF", details={"error": str(exc)}
        ) from exc
    return _finish(units, issues, preserve_empty=True)


def extract_docx(path: Path) -> ExtractionResult:
    from .office_structure import extract_structured_docx

    return extract_structured_docx(path)


def extract_pptx(path: Path) -> ExtractionResult:
    from .office_structure import extract_structured_pptx

    return extract_structured_pptx(path)


def _cell_value(cell) -> str:
    value = cell.value
    if value is None:
        return ""
    if isinstance(value, str):
        return normalize_text(value)
    return str(value)


_SPREADSHEETML_NAMESPACE = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_XLSX_STYLES_MEMBER = "xl/styles.xml"


def _xlsx_styles_without_invalid_font_families(styles: bytes) -> tuple[bytes, int]:
    """Remove only numeric font-family metadata outside OpenPyXL's safe range."""

    try:
        root = ElementTree.fromstring(styles)
    except (DefusedXmlException, ElementTree.ParseError) as exc:
        raise ExtractionError("could not parse XLSX styles XML") from exc
    fonts = root.find(f"{{{_SPREADSHEETML_NAMESPACE}}}fonts")
    if fonts is None:
        return styles, 0

    family_tag = f"{{{_SPREADSHEETML_NAMESPACE}}}family"
    removed = 0
    for font in list(fonts):
        for child in list(font):
            if child.tag != family_tag:
                continue
            raw_value = child.attrib.get("val")
            try:
                value = int(raw_value) if raw_value is not None else None
            except ValueError:
                continue
            if value is not None and not 0 <= value <= 14:
                font.remove(child)
                removed += 1
    if not removed:
        return styles, 0
    return ElementTree.tostring(
        root,
        encoding="utf-8",
        xml_declaration=True,
    ), removed


def _write_xlsx_with_safe_font_families(source: Path, destination: Path) -> int:
    """Write a temporary package with invalid font-family metadata omitted."""

    removed = 0
    with (
        zipfile.ZipFile(source) as package,
        zipfile.ZipFile(destination, "w", allowZip64=True) as normalized,
    ):
        normalized.comment = package.comment
        for member in package.infolist():
            if member.filename == _XLSX_STYLES_MEMBER:
                if member.file_size > MAX_XML_MEMBER_BYTES:
                    raise ExtractionError(
                        "XLSX styles XML exceeds the extraction limit",
                        details={
                            "member_bytes": member.file_size,
                            "limit": MAX_XML_MEMBER_BYTES,
                        },
                    )
                data = package.read(member)
                data, member_removed = _xlsx_styles_without_invalid_font_families(data)
                removed += member_removed
                normalized.writestr(member, data)
                continue
            with (
                package.open(member) as source_member,
                normalized.open(member, "w", force_zip64=True) as destination_member,
            ):
                shutil.copyfileobj(
                    source_member, destination_member, length=1024 * 1024
                )
    return removed


def _open_xlsx_workbook(load_workbook, path: Path):
    package = path.open("rb")
    try:
        workbook = load_workbook(
            filename=package,
            read_only=True,
            data_only=False,
            keep_links=False,
        )
    except Exception:
        package.close()
        raise
    return package, workbook


def extract_xlsx(path: Path) -> ExtractionResult:
    _preflight_zip(path)
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ExtractionError("openpyxl is not installed") from exc
    units: list[UnitDraft] = []
    issues: list[dict] = []
    cells_seen = 0
    package = None
    workbook = None
    temporary = None
    try:
        try:
            package, workbook = _open_xlsx_workbook(load_workbook, path)
        except ValueError:
            temporary = tempfile.TemporaryDirectory(prefix="corpus-xlsx-")
            normalized_path = Path(temporary.name) / "normalized.xlsx"
            removed = _write_xlsx_with_safe_font_families(path, normalized_path)
            if not removed:
                raise
            package, workbook = _open_xlsx_workbook(load_workbook, normalized_path)
            issues.append(
                {
                    "code": "xlsx_invalid_font_family_ignored",
                    "severity": "info",
                    "message": (
                        "Invalid XLSX font-family metadata was ignored in a temporary "
                        "read-only copy."
                    ),
                    "details": {"removed_elements": removed},
                }
            )
        try:
            for sheet in workbook.worksheets:
                for row_index, row in enumerate(sheet.iter_rows(), start=1):
                    if row_index > MAX_SHEET_ROWS or cells_seen >= MAX_SHEET_CELLS:
                        issues.append(
                            {
                                "code": "sheet_limit_reached",
                                "severity": "warning",
                                "message": (
                                    "Workbook extraction stopped at the configured cell limit."
                                ),
                                "sheet": sheet.title,
                            }
                        )
                        break
                    rendered: list[str] = []
                    first_cell = None
                    last_cell = None
                    for cell in row:
                        cells_seen += 1
                        value = _cell_value(cell)
                        if not value:
                            continue
                        first_cell = first_cell or cell.coordinate
                        last_cell = cell.coordinate
                        rendered.append(f"{cell.coordinate}={value}")
                    if rendered:
                        units.append(
                            UnitDraft(
                                "sheet_row",
                                {
                                    "sheet": sheet.title,
                                    "row": row_index,
                                    "range": f"{first_cell}:{last_cell}",
                                },
                                "\t".join(rendered),
                            )
                        )
        finally:
            workbook.close()
            package.close()
            workbook = None
            package = None
    except Exception as exc:
        raise ExtractionError(
            "could not extract XLSX", details={"error": str(exc)}
        ) from exc
    finally:
        if workbook is not None:
            workbook.close()
        if package is not None:
            package.close()
        if temporary is not None:
            temporary.cleanup()
    return _finish(units, issues)


EXTRACTORS = {
    "text": extract_text,
    "markdown": extract_markdown,
    "html": extract_html,
    "hwpx": extract_hwpx,
    "pdf": extract_pdf,
    "docx": extract_docx,
    "pptx": extract_pptx,
    "xlsx": extract_xlsx,
}


def extract(path: Path, adapter: str) -> ExtractionResult:
    extractor = EXTRACTORS.get(adapter)
    if extractor is None:
        raise ExtractionError(
            "no extractor is registered", details={"adapter": adapter}
        )
    return extractor(path)
